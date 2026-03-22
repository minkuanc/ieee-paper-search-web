"""
Writes/updates a papers.xlsx file with downloaded paper metadata.
Deduplicates by DOI so re-running won't create duplicate rows.
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelWriter:
    COLUMNS = [
        "Title",
        "Authors",
        "Year",
        "Venue",
        "DOI",
        "URL",
        "Local Path",
        "Download Status",
    ]

    def __init__(self, filepath: str):
        self.filepath = filepath
        if os.path.exists(filepath):
            self._wb = load_workbook(filepath)
            self._ws = self._wb.active
            self._existing_dois = self._load_existing_dois()
        else:
            self._wb = Workbook()
            self._ws = self._wb.active
            self._ws.title = "Papers"
            self._write_header()
            self._existing_dois = set()

    def _write_header(self):
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="00539B")  # IEEE blue
        for col_idx, col_name in enumerate(self.COLUMNS, start=1):
            cell = self._ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def _load_existing_dois(self) -> set:
        """Collect all DOIs already in the sheet (column 5 = DOI)."""
        dois = set()
        doi_col = self.COLUMNS.index("DOI") + 1
        for row in self._ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= doi_col:
                doi = row[doi_col - 1]
                if doi:
                    dois.add(str(doi).strip())
        return dois

    def append_papers(self, papers: list[dict]):
        """
        Append papers to the sheet, skipping any whose DOI already exists.

        Each dict in `papers` should have:
          title, authors, year, venue, doi, url, local_path, status
        """
        for paper in papers:
            doi = str(paper.get("doi", "")).strip()
            if doi and doi in self._existing_dois:
                continue  # deduplicate
            row = [
                paper.get("title", ""),
                paper.get("authors", ""),
                paper.get("year", ""),
                paper.get("venue", ""),
                doi,
                paper.get("url", ""),
                paper.get("local_path", ""),
                paper.get("status", ""),
            ]
            self._ws.append(row)
            if doi:
                self._existing_dois.add(doi)

    def save(self):
        """Auto-size columns and save the workbook."""
        for col_idx, col_name in enumerate(self.COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(col_name)
            for row in self._ws.iter_rows(
                min_row=2, min_col=col_idx, max_col=col_idx, values_only=True
            ):
                cell_val = str(row[0]) if row[0] is not None else ""
                max_len = max(max_len, len(cell_val))
            self._ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        self._wb.save(self.filepath)
